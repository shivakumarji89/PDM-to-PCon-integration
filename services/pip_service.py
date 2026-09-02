"""Parse a Product Information Page (PIP) sheet from the vocabulary workbook.

The workbook has one sheet per product (e.g. ``Rectangle, Post Leg``). Each
sheet lays the product out as ``Feature 1..N`` columns: a header row of feature
labels, a row of property names, then the values stacked below each property
with the order code in the next column. This service reads ONE such sheet into a
:class:`PipProduct` - the authoritative spec Class Creation is validated against.

Layout (verified against the Everywhere Tables vocabulary):
  * a "feature row" whose cells start with ``Feature `` marks the property block;
  * the NEXT row holds the property names, one per feature column;
  * each property's values sit in its own column and the code in the column to
    its right, read downward until both run out;
  * a property named ``.`` is the base/tail separator (no values);
  * a ``Notes`` column holds free-text engineering notes.
"""
from __future__ import annotations

import re
from pathlib import Path

from models.pip import PipDiff, PipDiffItem, PipProduct, PipProperty, PipValue
from services.base_service import BaseService


class PipService(BaseService):
    """Read one PIP product sheet into a :class:`PipProduct`."""

    def sheet_names(self, path: str | Path) -> list[str]:
        """The worksheet names in the workbook."""
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def parse_sheet(self, path: str | Path, sheet: str) -> PipProduct:
        """Parse one product sheet into a :class:`PipProduct`."""
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        return self._parse_rows(rows, sheet)

    # -- parsing --------------------------------------------------------

    @staticmethod
    def _text(cell) -> str:
        return "" if cell is None else str(cell).strip()

    def _parse_rows(self, rows: list[list], sheet: str) -> PipProduct:
        product = PipProduct(sheet=sheet)
        if not rows:
            return product

        feature_row = self._find_feature_row(rows)
        if feature_row is None:
            return product
        name_row = feature_row + 1
        if name_row >= len(rows):
            return product

        # Property columns come from the NAME row: (column, name) left-to-right.
        columns = [
            (col, self._text(cell))
            for col, cell in enumerate(rows[name_row])
            if self._text(cell)
        ]
        named_cols = {c for c, _n in columns}
        # The base/tail "." can sit in an UNNAMED column in the data (super-product
        # sheets) rather than the name row; pick it up from the first data row.
        if name_row + 1 < len(rows):
            for col, cell in enumerate(rows[name_row + 1]):
                if self._text(cell) == "." and col not in named_cols:
                    columns.append((col, "."))
        columns.sort(key=lambda cn: cn[0])
        value_cols = [c for c, n in columns if n not in (".",) and n.lower() != "notes"]
        notes_cols = {col for col, name in columns if name.lower() == "notes"}

        # Title: a real product title sits above the feature row; skip the
        # "Item prefix" label and property-name-like cells, else use the sheet.
        product.title = sheet
        for r in range(0, feature_row + 1):
            cand = next((self._text(c) for c in rows[r] if self._text(c)), "")
            if cand and cand.lower() not in ("item prefix",) and not cand.lower().startswith("feature "):
                if cand not in {n for _c, n in columns}:
                    product.title = cand
                    break

        # Bound the vocabulary block: it ends where the next block's header
        # begins (a repeated feature row, or the "Generic Product Prefix"/"Item
        # prefix" label) or at the first fully-empty value row.
        header_markers = {"generic product prefix", "item prefix"}
        block_end = len(rows)
        for r in range(name_row + 1, len(rows)):
            row = rows[r]
            texts = [self._text(c) for c in row]
            if sum(1 for t in texts if t.lower().startswith("feature ")) >= 2:
                block_end = r
                break
            if any(t.lower() in header_markers for t in texts):
                block_end = r
                break
            if all(
                not (self._text(row[c]) if c < len(row) else "")
                for c in value_cols
            ):
                block_end = r
                break

        order = 0
        for col, name in columns:
            if name.lower() == "notes":
                continue
            if name == ".":
                product.properties.append(
                    PipProperty(name=".", order=order, is_separator=True)
                )
                order += 1
                continue
            values: list[PipValue] = []
            for r in range(name_row + 1, block_end):
                row = rows[r]
                value = self._text(row[col]) if col < len(row) else ""
                code = self._text(row[col + 1]) if col + 1 < len(row) else ""
                if not value and not code:
                    continue
                values.append(PipValue(value=value, code=code))
            product.properties.append(
                PipProperty(name=name, order=order, values=values)
            )
            order += 1

        # Collect notes (a single Notes column running down the block).
        notes: list[str] = []
        for col in notes_cols:
            for r in range(name_row + 1, block_end):
                row = rows[r]
                text = self._text(row[col]) if col < len(row) else ""
                if text:
                    notes.append(text)
        if notes and product.properties:
            product.properties[-1].notes = notes
        return product

    def _find_feature_row(self, rows: list[list]) -> int | None:
        """Row index whose cells are the ``Feature 1..N`` labels (>=2 of them)."""
        for i, row in enumerate(rows):
            hits = sum(
                1 for c in row if self._text(c).lower().startswith("feature ")
            )
            if hits >= 2:
                return i
        return None

    # -- PDM reconstruction (PIP straight from the database) ------------

    def from_pdm(self, product_id) -> PipProduct:
        """Reconstruct a product's PIP directly from PDM, no Excel.

        Classifies each attribute functional (head/item-prefix, no
        ``OrderCodeFormatKey`` token) vs physical (tail, has a token), orders the
        head by ``DisplayOrder`` and the tail by ``ProductRange.OrderCodeFormatString``
        token order (options appended to the tail), inserts the ``.`` separator
        between, and attaches the product's application-text notes. The functional
        HEAD codes are not stored in PDM (positional in the SKU) so those values
        carry a name with an empty code.
        """
        repo = self.context.pdm_service.repository
        header = repo.fetch_pip_header(product_id)
        product = PipProduct()
        if header:
            code, rng, fmt = header[0][0], header[0][1], header[0][2]
            product.title = f"{code or ''} ({rng or ''})".strip()
            fmt = fmt or ""
        else:
            fmt = ""
        # {TOKEN} assembly order for the tail.
        token_order = {t: i for i, t in enumerate(re.findall(r"\{([^}]+)\}", fmt))}

        functional: list[tuple] = []  # (display_order, name, key, values)
        physical: list[tuple] = []    # (token_index, name, key, values)
        for aid, name, key, values in self._group_pip_rows(
            repo.fetch_pip_attributes(product_id)
        ):
            token = (key or "").strip("{}")
            coded = any(v.code for v in values)
            if token or coded:  # physical / tail
                idx = token_order.get(token, 10_000 + len(physical))
                physical.append((idx, name, key, values))
            else:  # functional / head (positional codes not stored)
                functional.append((len(functional), name, key, values))

        # Options are tail features too (they carry a format key + coded values).
        for oid, name, key, values in self._group_pip_rows(
            repo.fetch_pip_options(product_id)
        ):
            token = (key or "").strip("{}")
            idx = token_order.get(token, 10_000 + len(physical))
            physical.append((idx, name, key, values))

        order = 0
        for _do, name, _key, values in sorted(functional, key=lambda t: t[0]):
            product.properties.append(PipProperty(name=name, order=order, values=values))
            order += 1
        product.properties.append(PipProperty(name=".", order=order, is_separator=True))
        order += 1
        for _ti, name, _key, values in sorted(physical, key=lambda t: t[0]):
            product.properties.append(PipProperty(name=name, order=order, values=values))
            order += 1

        product.notes = [
            str(r[0]).strip() for r in repo.fetch_pip_notes(product_id) if r and r[0]
        ]
        return product

    @staticmethod
    def _group_pip_rows(rows):
        """Group flat (id, name, key, ..., value, code) PDM rows by entity into
        ``(id, name, key, [PipValue])`` preserving first-seen order.

        Attribute rows: (AttrId, AttrName, Type, Key, DisplayOrder, HDO, ValueId,
        ValueName, Code, Ordinal). Option rows: (OptId, OptName, Key, DisplayOrder,
        ValueId, ValueName, Code, Ordinal). The name is col1; the format key and
        value/code positions differ, so detect by length.
        """
        grouped: dict = {}
        order: list = []
        for r in rows:
            if len(r) >= 10:  # attribute row
                eid, name, key = r[0], r[1], r[3]
                vname, code = r[7], r[8]
            else:  # option row
                eid, name, key = r[0], r[1], r[2]
                vname, code = r[5], r[6]
            if eid not in grouped:
                grouped[eid] = (name, key, [])
                order.append(eid)
            value = "" if vname is None else str(vname).strip()
            codev = "" if code is None else str(code).strip()
            if value or codev:
                grouped[eid][2].append(PipValue(value=value, code=codev))
        return [(eid, grouped[eid][0], grouped[eid][1], grouped[eid][2]) for eid in order]

    # -- Class Creation view (the authored snapshot as a PIP) -----------

    def from_snapshot(self, snapshot) -> PipProduct:
        """Render the active Class Creation state as a :class:`PipProduct`.

        Mirrors :meth:`from_pdm`'s shape so the two can be diffed directly:
        functional (head, no ``OrderCodeFormatKey`` token and uncoded) properties
        first in ``DisplayOrder``, the ``.`` separator, then physical (tail)
        properties and options - all in ``DisplayOrder`` (the app's canonical
        order). Value codes are the ones the user AUTHORED in Class Creation.
        """
        product = PipProduct()
        if snapshot is None:
            return product
        prod = getattr(snapshot, "product", None)
        if prod is not None:
            product.title = f"{prod.code or ''} ({prod.range_name or ''})".strip()

        prop_values: dict[str, list] = {}
        for value in snapshot.property_values:
            prop_values.setdefault(str(value.property_id), []).append(value)
        opt_values: dict[str, list] = {}
        for value in snapshot.option_values:
            opt_values.setdefault(str(value.option_id), []).append(value)

        def _pip_values(raw: list) -> list[PipValue]:
            ordered = sorted(
                raw,
                key=lambda v: (v.display_order is None, v.display_order or 0, v.value or ""),
            )
            return [
                PipValue(value=(v.value or "").strip(), code=(v.code or "").strip())
                for v in ordered
            ]

        functional: list[tuple] = []
        physical: list[tuple] = []
        for prop in snapshot.properties:
            raw = prop.values or prop_values.get(str(prop.id), [])
            values = _pip_values(raw)
            # Head vs tail by OrderCodeFormatKey token ONLY - head values can be
            # authored with codes (the config-code decoder) yet stay functional.
            token = (prop.code or "").strip("{}")
            entry = (prop, values)
            (physical if token else functional).append(entry)
        # Options are always tail features, appended after the attributes.
        options: list[tuple] = [
            (option, _pip_values(option.values or opt_values.get(str(option.id), [])))
            for option in snapshot.options
        ]

        def _order_key(entry):
            ent = entry[0]
            do = getattr(ent, "display_order", None)
            return (do is None, do or 0, (ent.name or ""))

        order = 0
        for ent, values in sorted(functional, key=_order_key):
            product.properties.append(
                PipProperty(name=(ent.name or "").strip(), order=order, values=values)
            )
            order += 1
        product.properties.append(PipProperty(name=".", order=order, is_separator=True))
        order += 1
        for ent, values in sorted(physical, key=_order_key) + sorted(options, key=_order_key):
            product.properties.append(
                PipProperty(name=(ent.name or "").strip(), order=order, values=values)
            )
            order += 1
        return product

    # -- diff (PIP ground truth vs Class Creation) ----------------------

    @staticmethod
    def _norm(text: str) -> str:
        return (text or "").strip().lower()

    def _head_names(self, product: PipProduct) -> set[str]:
        """Normalized names of the head (pre-``.``) properties."""
        head: set[str] = set()
        for prop in product.properties:
            if prop.is_separator:
                break
            head.add(self._norm(prop.name))
        return head

    def diff(self, expected: PipProduct, actual: PipProduct) -> PipDiff:
        """Compare a ground-truth PIP (``expected``) against Class Creation
        (``actual``) and return a :class:`PipDiff`.

        Checks property presence, the head/tail (``.``) split, the value set per
        property, and TAIL value codes. HEAD value codes are positional in the
        SKU (not stored in PDM) so they are reported as decoder-pending, not
        compared. Property order within the head (both ``DisplayOrder``) is a
        warning; tail order is not asserted (display vs code-assembly order).
        """
        norm = self._norm
        result = PipDiff(title=expected.title or actual.title)

        exp_props = [p for p in expected.properties if not p.is_separator]
        act_props = [p for p in actual.properties if not p.is_separator]
        exp_by = {norm(p.name): p for p in exp_props}
        act_by = {norm(p.name): p for p in act_props}
        exp_head = self._head_names(expected)
        act_head = self._head_names(actual)

        for name, prop in exp_by.items():
            if name not in act_by:
                result.items.append(PipDiffItem(
                    "error", "missing_property",
                    f"Property '{prop.name}' is in the PIP but not in Class Creation.",
                ))
        for name, prop in act_by.items():
            if name not in exp_by:
                result.items.append(PipDiffItem(
                    "warning", "extra_property",
                    f"Property '{prop.name}' is in Class Creation but not in the PIP.",
                ))

        head_pending = 0
        for name, ep in exp_by.items():
            ap = act_by.get(name)
            if ap is None:
                continue
            is_head = name in exp_head
            if is_head != (name in act_head):
                result.items.append(PipDiffItem(
                    "error", "split",
                    f"Property '{ep.name}' is "
                    f"{'head/functional' if is_head else 'tail/physical'} in the PIP "
                    f"but {'head' if name in act_head else 'tail'} in Class Creation.",
                ))
            exp_vals = {norm(v.value): v for v in ep.values}
            act_vals = {norm(v.value): v for v in ap.values}
            for vn, ev in exp_vals.items():
                av = act_vals.get(vn)
                if av is None:
                    result.items.append(PipDiffItem(
                        "error", "missing_value",
                        f"'{ep.name}' / '{ev.value}': in the PIP but not in Class Creation.",
                    ))
                    continue
                exp_code = (ev.code or "").strip()
                act_code = (av.code or "").strip()
                # HEAD codes are positional in the SKU: PDM leaves them blank
                # (unverifiable -> pending), but an Excel PIP prints them, so when
                # the source supplies a head code we DO validate the authored one.
                if is_head and not exp_code:
                    head_pending += 1
                    continue
                if exp_code and act_code and exp_code != act_code:
                    result.items.append(PipDiffItem(
                        "error", "code_mismatch",
                        f"'{ep.name}' / '{ev.value}': PIP code '{exp_code}' but "
                        f"Class Creation code '{act_code}'.",
                    ))
                elif exp_code and not act_code:
                    result.items.append(PipDiffItem(
                        "warning", "missing_code",
                        f"'{ep.name}' / '{ev.value}': PIP code '{exp_code}' but "
                        f"Class Creation value is uncoded.",
                    ))
            for vn, av in act_vals.items():
                if vn not in exp_vals:
                    result.items.append(PipDiffItem(
                        "warning", "extra_value",
                        f"'{ap.name}' / '{av.value}': in Class Creation but not in the PIP.",
                    ))

        # Head order (both DisplayOrder) - matched head names in each side's order.
        exp_head_seq = [norm(p.name) for p in exp_props if norm(p.name) in exp_head
                        and norm(p.name) in act_by]
        act_head_seq = [norm(p.name) for p in act_props if norm(p.name) in act_head
                        and norm(p.name) in exp_by]
        if exp_head_seq != act_head_seq:
            result.items.append(PipDiffItem(
                "warning", "order",
                "Head (functional) property order differs from the PIP.",
            ))

        if head_pending:
            result.items.append(PipDiffItem(
                "info", "head_pending",
                f"{head_pending} head (functional) value code(s) are positional in "
                "the SKU and not verified here (decoder-pending).",
            ))
        return result

    def validate_class_creation(self, product_id, snapshot=None) -> PipDiff:
        """Diff the PDM-reconstructed PIP for ``product_id`` against the active
        Class Creation snapshot. Convenience wrapper over
        :meth:`from_pdm` + :meth:`from_snapshot` + :meth:`diff`."""
        snap = snapshot if snapshot is not None else self.context.active_snapshot
        return self.diff(self.from_pdm(product_id), self.from_snapshot(snap))

    def validate_class_creation_excel(self, path, sheet, snapshot=None) -> PipDiff:
        """Diff an Excel PIP sheet against the active Class Creation snapshot.

        Unlike :meth:`validate_class_creation` (PDM), an Excel PIP prints the
        functional HEAD order codes, so those are validated too (not reported as
        decoder-pending). Convenience over :meth:`parse_sheet` + :meth:`diff`."""
        snap = snapshot if snapshot is not None else self.context.active_snapshot
        return self.diff(self.parse_sheet(path, sheet), self.from_snapshot(snap))

